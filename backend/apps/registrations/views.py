from rest_framework import status
from rest_framework.generics import (
    CreateAPIView,
    RetrieveAPIView,
)

from rest_framework.permissions import (
    AllowAny,
)

from rest_framework.response import Response
from rest_framework.views import APIView

from apps.events.models import Event

from .models import Registration
from .serializers import (
    PublicRegistrationSerializer,
    RegistrationFormSerializer,
)
from .services import create_registration


class PublicRegistrationFormView(
    RetrieveAPIView
):

    permission_classes = [
        AllowAny,
    ]

    serializer_class = (
        RegistrationFormSerializer
    )

    lookup_url_kwarg = "event_id"

    def get_object(self):

        event = Event.objects.get(
            id=self.kwargs["event_id"],
            is_active=True,
        )

        return event.registration_form


class PublicRegistrationCreateView(
    APIView
):

    permission_classes = [
        AllowAny,
    ]

    def post(
        self,
        request,
        event_id,
    ):

        try:

            event = Event.objects.get(
                id=event_id,
                is_active=True,
            )

        except Event.DoesNotExist:

            return Response(
                {
                    "detail": (
                        "Event not found."
                    )
                },
                status=status.HTTP_404_NOT_FOUND,
            )

        serializer = (
            PublicRegistrationSerializer(
                data=request.data,
                context={
                    "event": event,
                },
            )
        )

        serializer.is_valid(
            raise_exception=True
        )

        registration = (
            create_registration(
                event=event,
                category=serializer.validated_data[
                    "category"
                ],
                participant_data=(
                    serializer.validated_data[
                        "participant"
                    ]
                ),
                form_data=(
                    serializer.validated_data[
                        "form_data"
                    ]
                ),
                reserve=serializer.validated_data.get(
                    "reserve", False
                ),
            )
        )

        return Response(
            {
                "registration": {
                    "id": registration.id,
                    "registration_number": (
                        registration
                        .registration_number
                    ),
                    "status": (
                        registration.status
                    ),
                    "amount": str(
                        registration.amount
                    ),
                    "currency": (
                        registration.currency
                    ),
                }
            },
            status=status.HTTP_201_CREATED,
        )


class PublicRegistrationLookupView(APIView):
    """
    GET /api/v1/registrations/public/lookup/?q=<reference-or-email>

    Powers the "track your registration" feature on the public site.
    Deliberately returns a small, non-sensitive subset of fields.
    """

    permission_classes = [AllowAny]

    def get(self, request):
        query = request.query_params.get("q", "").strip()

        if not query:
            return Response(
                {"detail": "Provide a reference number or email as ?q="},
                status=status.HTTP_400_BAD_REQUEST,
            )

        registration = (
            Registration.objects
            .select_related("participant", "category")
            .filter(registration_number__iexact=query)
            .first()
            or Registration.objects
            .select_related("participant", "category")
            .filter(participant__email__iexact=query)
            .order_by("-registered_at")
            .first()
        )

        if not registration:
            return Response(
                {"detail": "No matching registration found."},
                status=status.HTTP_404_NOT_FOUND,
            )

        return Response(
            {
                "reference": registration.registration_number,
                "status": registration.status,
                "category": registration.category.name,
                "amount": str(registration.amount),
                "currency": registration.currency,
                "full_name": (
                    f"{registration.participant.first_name} "
                    f"{registration.participant.last_name}"
                ).strip(),
                "submitted_at": registration.registered_at,
            }
        )

# ---------------------------------------------------------------------------
# Admin-facing views
# ---------------------------------------------------------------------------

import csv
import io
import re

import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils.dateparse import parse_datetime

from rest_framework import filters
from rest_framework.generics import ListAPIView, RetrieveUpdateDestroyAPIView
from rest_framework.parsers import JSONParser, MultiPartParser
from rest_framework.permissions import IsAuthenticated
from django_filters.rest_framework import DjangoFilterBackend

from apps.common.access import (
    EVENT_REGISTRATION_MANAGE_ROLES,
    EVENT_VIEW_ROLES,
    require_event_role,
)
from apps.common.permissions import HasEventRole

from .models import RegistrationCategory
from .serializers import (
    AdminManualRegistrationSerializer,
    AdminRegistrationSerializer,
    AdminRegistrationStatusUpdateSerializer,
)
from .services import create_registration, generate_registration_number


class AdminRegistrationListView(ListAPIView):
    """
    GET /api/v1/registrations/admin/events/<event_id>/registrations/

    Supports ?status=, ?category=, ?search=, ?ordering= and standard
    pagination — this is the data source for the admin registrations
    table. ?gender=, ?organisation=, ?attendance_type= filter on the
    equivalent keys inside form_data (there's no dedicated column for
    these — the registration form is admin-configurable per event, so
    these answers live in the JSON blob rather than fixed model fields).
    """

    permission_classes = [IsAuthenticated, HasEventRole(*EVENT_VIEW_ROLES)]
    serializer_class = AdminRegistrationSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ["status", "category"]
    search_fields = [
        "registration_number",
        "participant__first_name",
        "participant__last_name",
        "participant__email",
        "participant__phone",
    ]
    ordering_fields = ["registered_at", "amount", "status"]

    def get_queryset(self):
        qs = (
            Registration.objects
            .select_related("participant", "category", "event")
            .filter(event_id=self.kwargs["event_id"])
        )

        gender = self.request.query_params.get("gender")
        if gender:
            qs = qs.filter(form_data__gender=gender)

        organisation = self.request.query_params.get("organisation")
        if organisation:
            qs = qs.filter(form_data__club_or_institution=organisation)

        attendance_type = self.request.query_params.get("attendance_type")
        if attendance_type:
            qs = qs.filter(form_data__attendance_type=attendance_type)

        return qs


class AdminRegistrationFilterOptionsView(APIView):
    """
    GET /api/v1/registrations/admin/events/<event_id>/registrations/filters/

    Distinct values for the admin table's filter dropdowns — genders/
    organisations/attendance types come from form_data (see
    AdminRegistrationListView), so they can't be hardcoded; categories are
    event-specific too. Small and cheap enough to just compute directly
    rather than caching.
    """

    permission_classes = [IsAuthenticated, HasEventRole(*EVENT_VIEW_ROLES)]

    def get(self, request, event_id):
        qs = Registration.objects.filter(event_id=event_id)

        def distinct(key):
            values = qs.values_list(f"form_data__{key}", flat=True).distinct()
            return sorted({v for v in values if v})

        # code/price/currency included so the admin's bulk-upload feature
        # can build its downloadable template and category-code guide from
        # this same call, rather than needing a second endpoint.
        categories = list(
            RegistrationCategory.objects
            .filter(event_id=event_id)
            .values("id", "name", "code", "price", "currency")
            .order_by("name")
        )

        return Response(
            {
                "categories": categories,
                "genders": distinct("gender"),
                "organisations": distinct("club_or_institution"),
                "attendance_types": distinct("attendance_type"),
            }
        )


class AdminRegistrationDetailView(RetrieveUpdateDestroyAPIView):
    """
    GET/PATCH/DELETE /api/v1/registrations/admin/registrations/<id>/

    PATCH only accepts {"status": "..."} — use this to flip a
    registration from RESERVED/PENDING_PAYMENT to CONFIRMED once a
    manual/cash payment has been taken, or to CANCELLED/REFUNDED.
    """

    permission_classes = [IsAuthenticated]
    serializer_class = AdminRegistrationSerializer
    queryset = Registration.objects.select_related("participant", "category", "event")

    def get_object(self):
        # Not keyed by event_id in the URL (it's looked up by
        # registration id), so the role check happens here once the
        # event is known, rather than via a HasEventRole permission
        # class. GET only needs view access; PATCH/DELETE (both of which
        # call this via the generic dispatch) need registration-manage
        # access.
        registration = super().get_object()

        required_roles = (
            EVENT_VIEW_ROLES
            if self.request.method == "GET"
            else EVENT_REGISTRATION_MANAGE_ROLES
        )

        require_event_role(
            self.request.user,
            registration.event_id,
            *required_roles,
        )

        return registration

    def patch(self, request, *args, **kwargs):
        registration = self.get_object()

        serializer = AdminRegistrationStatusUpdateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        new_status = serializer.validated_data["status"]
        old_status = registration.status

        registration.status = new_status
        registration.save(update_fields=["status", "updated_at"])

        # Manually marking something CONFIRMED (e.g. cash paid at the
        # door) should notify the runner just like an online payment
        # would, but only if we're actually transitioning INTO confirmed.
        if new_status == Registration.Status.CONFIRMED and old_status != new_status:
            from apps.notifications.services import notify_payment_confirmed
            notify_payment_confirmed(registration)

        return Response(AdminRegistrationSerializer(registration).data)

    def perform_destroy(self, instance):
        # Payment.registration is on_delete=PROTECT, so instance.delete()
        # below would 500 with an unhandled ProtectedError for almost any
        # registration that ever attempted a payment — which is most of
        # them. Refuse to delete one with a real successful payment (that
        # would destroy actual revenue history — cancel/refund it
        # instead), but clear out failed/abandoned payment attempts so
        # deleting an abandoned or duplicate registration actually works.
        from rest_framework.exceptions import ValidationError

        from apps.payments.models import Payment

        if Payment.objects.filter(
            registration=instance, status=Payment.Status.SUCCESS
        ).exists():
            raise ValidationError(
                "This registration has a successful payment and can't be "
                "deleted — cancel or refund it instead."
            )

        Payment.objects.filter(registration=instance).delete()
        instance.delete()


class AdminRegistrationEditView(APIView):
    """
    PATCH /api/v1/registrations/admin/registrations/<id>/details/

    Lets an admin correct participant/form_data fields on an existing
    registration — same field set as the manual "Add person" form and
    bulk upload, minus email: deliberately not editable here. Once a
    registration exists, its email is where every notification for it
    has gone and will go; letting it be changed here could quietly
    redirect a paid registration to a different inbox.

    Separate from AdminRegistrationDetailView.patch() (which only
    changes status) so status-change logic — including the "you're
    confirmed" notification — is untouched by this endpoint.
    """

    permission_classes = [IsAuthenticated]

    PARTICIPANT_FIELDS = ["first_name", "last_name", "phone"]
    FORM_DATA_FIELDS = [
        "gender", "age_range", "country", "tshirt_size", "attendance_type",
        "club_or_institution", "emergency_contact_name", "emergency_contact_phone",
        "medical_notes",
    ]

    def patch(self, request, pk):
        registration = get_object_or_404(
            Registration.objects.select_related("participant", "category", "event"),
            pk=pk,
        )
        require_event_role(
            request.user, registration.event_id, *EVENT_REGISTRATION_MANAGE_ROLES
        )

        if "email" in request.data:
            return Response(
                {"detail": "Email address cannot be changed."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        participant_updates = {
            f: (request.data[f] or "").strip()
            for f in self.PARTICIPANT_FIELDS
            if f in request.data
        }

        for name_field in ("first_name", "last_name"):
            if name_field in participant_updates and not participant_updates[name_field]:
                return Response(
                    {"detail": f"{name_field.replace('_', ' ').title()} can't be blank."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        if participant_updates:
            for field, value in participant_updates.items():
                setattr(registration.participant, field, value)
            registration.participant.save(
                update_fields=[*participant_updates.keys(), "updated_at"]
            )

        form_data_updates = {
            f: request.data[f] for f in self.FORM_DATA_FIELDS if f in request.data
        }
        if form_data_updates:
            registration.form_data = {**registration.form_data, **form_data_updates}
            registration.save(update_fields=["form_data", "updated_at"])

        return Response(AdminRegistrationSerializer(registration).data)


class AdminRegistrationCreateView(APIView):
    """
    POST /api/v1/registrations/admin/events/<event_id>/registrations/

    Manual "register a participant" — used by the admin for walk-ins,
    phone registrations, etc. Defaults to CONFIRMED status (cash taken in
    person) but the admin can pick any status.
    """

    permission_classes = [
        IsAuthenticated,
        HasEventRole(*EVENT_REGISTRATION_MANAGE_ROLES),
    ]

    def post(self, request, event_id):
        event = Event.objects.get(id=event_id)

        serializer = AdminManualRegistrationSerializer(
            data=request.data, context={"event": event}
        )
        serializer.is_valid(raise_exception=True)

        category = serializer.validated_data["category_id"]

        registration = create_registration(
            event=event,
            category=category,
            participant_data=serializer.validated_data["participant"],
            form_data=serializer.validated_data.get("form_data", {}),
            reserve=False,
        )

        desired_status = serializer.validated_data["status"]
        if desired_status != registration.status:
            registration.status = desired_status
            registration.save(update_fields=["status", "updated_at"])

        return Response(
            AdminRegistrationSerializer(registration).data,
            status=status.HTTP_201_CREATED,
        )


class AdminRegistrationBulkUploadView(APIView):
    """
    POST /api/v1/registrations/admin/events/<event_id>/registrations/bulk-upload/

    Accepts EITHER a CSV/XLSX file (multipart field name "file") OR a JSON
    body {"rows": [...]} — the latter is what the admin's bulk-upload
    review screen sends once the admin has previewed/edited rows (see
    AdminRegistrationBulkUploadPreviewView below); both paths run through
    the exact same row-by-row validation and creation logic here, so the
    preview can never say a row is fine when the real commit would then
    reject it.

    Row columns: first_name, last_name, email, phone, category_code,
    status (optional, defaults to CONFIRMED), plus the same optional
    extra fields the admin's manual "Add person" form collects — gender,
    age_range, country, tshirt_size, attendance_type, club_or_institution,
    emergency_contact_name, emergency_contact_phone, medical_notes — all
    stored in form_data exactly like a manual registration would.

    Returns a report of created rows and any rows that failed, rather
    than failing the whole batch on one bad row.
    """

    permission_classes = [
        IsAuthenticated,
        HasEventRole(*EVENT_REGISTRATION_MANAGE_ROLES),
    ]
    parser_classes = [MultiPartParser, JSONParser]

    REQUIRED_COLUMNS = ["first_name", "last_name", "category_code"]
    FORM_DATA_COLUMNS = [
        "gender",
        "age_range",
        "country",
        "tshirt_size",
        "attendance_type",
        "club_or_institution",
        "emergency_contact_name",
        "emergency_contact_phone",
        "medical_notes",
    ]
    # Only the fields with a fixed set of valid values get a (non-blocking)
    # warning on mismatch — the rest are free text.
    KNOWN_VALUES = {
        "gender": {"male", "female"},
        "age_range": {"Under 18", "18-29", "30-39", "40-49", "50-59", "60+"},
        "tshirt_size": {"XS", "S", "M", "L", "XL", "XXL", "3XL", "4XL", "5XL"},
        "attendance_type": {"in-person", "virtual"},
    }
    EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

    def post(self, request, event_id):
        event = Event.objects.get(id=event_id)
        upload = request.FILES.get("file")

        if upload:
            rows = self._parse_rows(upload)
        else:
            rows = request.data.get("rows")
            if not isinstance(rows, list):
                return Response(
                    {"detail": "Attach a file under the 'file' field, or POST {'rows': [...]}."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        report = self._process_rows(event, rows, commit=True)

        return Response(
            report,
            status=status.HTTP_201_CREATED if report["created_count"] else status.HTTP_400_BAD_REQUEST,
        )

    def _process_rows(self, event, rows, commit):
        """
        Shared by the real upload (commit=True, actually creates
        registrations) and the preview endpoint (commit=False, only
        reports what *would* happen) so the two can never disagree.
        """
        categories = {
            c.code: c
            for c in RegistrationCategory.objects.filter(event=event)
        }

        created = []
        errors = []
        results = []
        seen_emails = {}

        for index, raw_row in enumerate(rows, start=2):  # header is row 1
            row = {
                (k or "").strip().lower(): ("" if v is None else str(v).strip())
                for k, v in (raw_row or {}).items()
            }

            row_errors = []
            row_warnings = []

            missing = [c for c in self.REQUIRED_COLUMNS if not row.get(c)]
            if missing:
                row_errors.append(f"Missing required field(s): {', '.join(missing)}")

            category = None
            code = row.get("category_code", "")
            if code:
                category = categories.get(code)
                if not category:
                    row_errors.append(f"Unknown category_code '{code}'")

            desired_status = (row.get("status") or "CONFIRMED").upper()
            if desired_status not in Registration.Status.values:
                row_errors.append(f"Unknown status '{desired_status}'")

            email = row.get("email", "")
            if email:
                if not self.EMAIL_RE.match(email):
                    row_warnings.append(f"'{email}' doesn't look like a valid email")
                elif email.lower() in seen_emails:
                    row_warnings.append(f"Duplicate email — also row {seen_emails[email.lower()]}")
                else:
                    seen_emails[email.lower()] = index

            for field, known in self.KNOWN_VALUES.items():
                value = row.get(field, "")
                if value and value not in known:
                    row_warnings.append(
                        f"'{value}' isn't one of the usual {field} values ({', '.join(sorted(known))}) — check spelling/casing"
                    )

            if row_errors:
                errors.append({"row": index, "error": "; ".join(row_errors)})
                results.append(
                    {"row": index, "valid": False, "errors": row_errors, "warnings": row_warnings, "data": row}
                )
                continue

            if not commit:
                results.append(
                    {"row": index, "valid": True, "errors": [], "warnings": row_warnings, "data": row}
                )
                continue

            form_data = {
                field: row[field] for field in self.FORM_DATA_COLUMNS if row.get(field)
            }

            try:
                registration = create_registration(
                    event=event,
                    category=category,
                    participant_data={
                        "first_name": row["first_name"],
                        "last_name": row["last_name"],
                        "email": row.get("email", ""),
                        "phone": row.get("phone", ""),
                    },
                    form_data=form_data,
                    reserve=False,
                )

                old_status = registration.status
                if desired_status != registration.status:
                    registration.status = desired_status
                    registration.save(update_fields=["status", "updated_at"])

                # create_registration() already sent the right notification
                # for whatever status it computed on its own (a "pending
                # payment" SMS, or — for an already-free category — the
                # confirmation email+SMS below). But bulk upload can then
                # override that status afterwards (e.g. importing rows
                # that were already paid in cash), and a plain .save()
                # doesn't know to notify anyone. Send the same "you're
                # confirmed" email+SMS the public site sends on a real
                # payment, exactly like AdminRegistrationDetailView.patch()
                # already does for the same kind of manual confirmation.
                if desired_status == Registration.Status.CONFIRMED and old_status != desired_status:
                    from apps.notifications.services import notify_payment_confirmed

                    notify_payment_confirmed(registration)

                created.append(registration.registration_number)
                results.append(
                    {
                        "row": index,
                        "valid": True,
                        "errors": [],
                        "warnings": row_warnings,
                        "reference": registration.registration_number,
                    }
                )

            except Exception as exc:  # noqa: BLE001
                errors.append({"row": index, "error": str(exc)})
                results.append({"row": index, "valid": False, "errors": [str(exc)], "warnings": row_warnings, "data": row})

        return {
            "created_count": len(created),
            "created_references": created,
            "error_count": len(errors),
            "errors": errors,
            "results": results,
        }

    def _parse_rows(self, upload):
        filename = (upload.name or "").lower()

        if filename.endswith(".csv"):
            text = upload.read().decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text))
            return [
                {(k or "").strip().lower(): (v or "").strip() for k, v in row.items()}
                for row in reader
            ]

        # Assume Excel otherwise
        workbook = openpyxl.load_workbook(upload, data_only=True)
        sheet = workbook.active

        rows_iter = sheet.iter_rows(values_only=True)
        headers = [str(h or "").strip().lower() for h in next(rows_iter)]

        rows = []
        for values in rows_iter:
            if all(v in (None, "") for v in values):
                continue
            row = {
                headers[i]: ("" if v is None else str(v).strip())
                for i, v in enumerate(values)
                if i < len(headers)
            }
            rows.append(row)

        return rows


class AdminRegistrationBulkUploadPreviewView(AdminRegistrationBulkUploadView):
    """
    POST /api/v1/registrations/admin/events/<event_id>/registrations/bulk-upload/preview/

    Same file parsing + validation as the real bulk upload (inherits its
    permission check and REQUIRED_COLUMNS), but creates nothing — it's a
    dry run so the admin can review and fix rows in the browser first.
    Always takes a file (the very first step, before any editing exists
    to send back as rows).
    """

    def post(self, request, event_id):
        event = Event.objects.get(id=event_id)
        upload = request.FILES.get("file")

        if not upload:
            return Response(
                {"detail": "Attach a file under the 'file' field."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        rows = self._parse_rows(upload)
        report = self._process_rows(event, rows, commit=False)

        return Response(report)


class AdminRegistrationBulkUploadTemplateView(APIView):
    """
    GET /api/v1/registrations/admin/events/<event_id>/registrations/bulk-upload/template/

    A ready-to-fill .xlsx for the bulk-upload feature: the exact header
    row AdminRegistrationBulkUploadView expects, two worked examples using
    this event's real category codes, and a second sheet listing every
    valid category code/name/price so the admin isn't guessing at codes.
    """

    permission_classes = [
        IsAuthenticated,
        HasEventRole(*EVENT_REGISTRATION_MANAGE_ROLES),
    ]

    # Same core columns as before, plus every optional field the admin's
    # manual "Add person" form collects — same field set, same order.
    COLUMNS = [
        "first_name", "last_name", "email", "phone", "category_code", "status",
        "gender", "age_range", "country", "tshirt_size", "attendance_type",
        "club_or_institution", "emergency_contact_name", "emergency_contact_phone",
        "medical_notes",
    ]

    def get(self, request, event_id):
        event = Event.objects.get(id=event_id)
        categories = list(
            RegistrationCategory.objects.filter(event=event).order_by("name")
        )

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Registrations"

        for col_index, header in enumerate(self.COLUMNS, start=1):
            sheet.cell(row=1, column=col_index, value=header)

        example_code = categories[0].code if categories else ""
        examples = [
            [
                "Jane", "Mwansa", "jane.mwansa@example.com", "0977000000", example_code, "CONFIRMED",
                "female", "30-39", "Zambia", "M", "in-person",
                "Copperbelt Runners Club", "John Mwansa", "0977111222", "",
            ],
            [
                "John", "Banda", "", "0966000000", example_code, "PENDING_PAYMENT",
                "male", "18-29", "Zambia", "L", "in-person",
                "", "", "", "",
            ],
        ]
        for row_index, example in enumerate(examples, start=2):
            for col_index, value in enumerate(example, start=1):
                sheet.cell(row=row_index, column=col_index, value=value)

        for col_index in range(1, len(self.COLUMNS) + 1):
            sheet.column_dimensions[get_column_letter(col_index)].width = 22

        codes_sheet = workbook.create_sheet("Category Codes")
        codes_sheet.append(["Category", "Code", "Price", "Currency"])
        for category in categories:
            codes_sheet.append(
                [category.name, category.code, float(category.price), category.currency]
            )
        for col_index in range(1, 5):
            codes_sheet.column_dimensions[get_column_letter(col_index)].width = 26

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{event.slug}-bulk-upload-template.xlsx"'
        )
        return response


class AdminRegistrationExportView(APIView):
    """
    GET /api/v1/registrations/admin/events/<event_id>/registrations/export/

    Streams an .xlsx of every registration for the event — this is what
    the admin's "Export to Excel" button hits directly (it's a normal
    GET, so the frontend can just set window.location or an <a href>).
    """

    permission_classes = [IsAuthenticated, HasEventRole(*EVENT_VIEW_ROLES)]

    # A mix of runner-form and vendor-form fields (both live in the same
    # form_data JSON blob, just under different keys per event) — using
    # .get() with a blank default means each event's registrations just
    # leave the other event's columns empty rather than needing a
    # per-event column configuration.
    COLUMNS = [
        ("Reference", lambda r: r.registration_number),
        ("Status", lambda r: r.get_status_display()),
        ("First name", lambda r: r.participant.first_name),
        ("Last name", lambda r: r.participant.last_name),
        ("Email", lambda r: r.participant.email),
        ("Phone", lambda r: r.participant.phone),
        ("Category", lambda r: r.category.name),
        ("Gender", lambda r: r.form_data.get("gender", "")),
        ("Age range", lambda r: r.form_data.get("age_range", "")),
        ("Country", lambda r: r.form_data.get("country", "")),
        ("T-shirt size", lambda r: r.form_data.get("tshirt_size", "")),
        ("Attendance", lambda r: r.form_data.get("attendance_type", "")),
        ("Organisation", lambda r: r.form_data.get("club_or_institution", "")),
        ("Emergency contact name", lambda r: r.form_data.get("emergency_contact_name", "")),
        ("Emergency contact phone", lambda r: r.form_data.get("emergency_contact_phone", "")),
        ("Medical notes", lambda r: r.form_data.get("medical_notes", "")),
        ("Business name", lambda r: r.form_data.get("business_name", "")),
        ("Business location", lambda r: r.form_data.get("business_location", "")),
        ("Products / services", lambda r: r.form_data.get("products_services", "")),
        ("Exhibition / activation requirement", lambda r: r.form_data.get("requirement", "")),
        ("Amount", lambda r: float(r.amount)),
        ("Currency", lambda r: r.currency),
        ("Registered at", lambda r: r.registered_at.replace(tzinfo=None) if r.registered_at else None),
    ]

    def get(self, request, event_id):
        event = Event.objects.get(id=event_id)

        registrations = (
            Registration.objects
            .select_related("participant", "category")
            .filter(event=event)
            .order_by("-registered_at")
        )

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Registrations"

        for col_index, (header, _) in enumerate(self.COLUMNS, start=1):
            sheet.cell(row=1, column=col_index, value=header)

        for row_index, registration in enumerate(registrations, start=2):
            for col_index, (_, getter) in enumerate(self.COLUMNS, start=1):
                sheet.cell(row=row_index, column=col_index, value=getter(registration))

        for col_index in range(1, len(self.COLUMNS) + 1):
            sheet.column_dimensions[get_column_letter(col_index)].width = 22

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{event.slug}-registrations.xlsx"'
        )
        return response
